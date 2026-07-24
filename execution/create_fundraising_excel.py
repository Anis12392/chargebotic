import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NAVY       = "1A252F"
BLUE       = "2C3E50"
LIGHT_BLUE = "2980B9"
GREEN      = "27AE60"
LIGHT_GREEN= "D5F5E3"
RED        = "C0392B"
ORANGE     = "E67E22"
YELLOW     = "FEF9E7"
WHITE      = "FFFFFF"
LIGHT_GRAY = "F2F2F2"
DARK_GRAY  = "7F8C8D"
PINK       = "FDEDEC"
PURPLE     = "6C3483"
LIGHT_PURPLE="F4ECF7"

def hdr(cell, bg=BLUE, fg=WHITE, bold=True, size=11, align="center"):
    cell.font = Font(bold=bold, color=fg, size=size, name="Calibri")
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)

def cel(cell, bold=False, color="000000", bg=None, align="left", size=10):
    cell.font = Font(bold=bold, color=color, size=size, name="Calibri")
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)

def bdr(cell):
    thin = Side(style="thin", color="CCCCCC")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

def cw(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def rh(ws, row, height):
    ws.row_dimensions[row].height = height

def merge_hdr(ws, r, c1, c2, text, bg=LIGHT_BLUE, fg=WHITE, size=11):
    ws.merge_cells(f"{get_column_letter(c1)}{r}:{get_column_letter(c2)}{r}")
    cell = ws.cell(row=r, column=c1, value=text)
    hdr(cell, bg, fg, True, size, "left")
    rh(ws, r, 22)

wb = openpyxl.Workbook()

# ══════════════════════════════════════════════════════════
# SHEET 1 — VISION & ROADMAP
# ══════════════════════════════════════════════════════════
ws0 = wb.active
ws0.title = "Vision & Roadmap"

ws0.merge_cells("A1:H1")
ws0["A1"] = "WIREBIRD AEROSPACE — COMPANY VISION & FUNDRAISING ROADMAP"
hdr(ws0["A1"], NAVY, WHITE, True, 16)
rh(ws0, 1, 50)

ws0.merge_cells("A2:H2")
ws0["A2"] = "Drone-based energy distribution from power lines → ground systems · Dual-use: Defense + Utilities · HQ: Los Angeles, CA"
hdr(ws0["A2"], DARK_GRAY, WHITE, False, 10)
rh(ws0, 2, 20)

ws0.merge_cells("A4:H4")
ws0["A4"] = "🎯  NORTH STAR: $1B REVENUE BY 2030 · Long-term: Energy distribution for Moon & Mars"
hdr(ws0["A4"], RED, WHITE, True, 13)
rh(ws0, 4, 35)

# Fundraising rounds table
merge_hdr(ws0, 6, 1, 8, "FUNDRAISING ROADMAP", LIGHT_BLUE, WHITE, 12)
for i, h_ in enumerate(["Round","Amount","Valuation","Timeline","Purpose","Key Milestones","Status",""],1):
    c = ws0.cell(row=7, column=i, value=h_)
    hdr(c, BLUE, WHITE)
rh(ws0, 7, 28)

rounds = [
    ["Angel SAFE",  "$20K",        "$5M cap",     "This week",    "Immediate bridge",           "—",                                                                "🔴 To do"],
    ["Bridge",      "$200K",       "$5M cap",     "July 2026",    "Runway to Pre-Seed",         "Angels signed · ADAM $80K pilot covers most of gap",               "🔴 Launch"],
    ["Pre-Seed",    "$2M",         "$10–15M",     "Oct 2026",     "Production + hangar",        "ADAM deal signed · JIFX demo · Patent · SBIR submitted (Aug)",     "🎯 Oct 2026"],
    ["SBIR DoD Ph.I","$150–300K",   "Non-dilutive","~Mar 2027",    "DoD validation + cash",      "Submit Aug 2026 → ~9 months → award ~March 2027",                  "📅 Submit Aug"],
    ["SBIR SOCOM Ph.I","$150–300K","Non-dilutive","~Mar 2027",    "SOCOM validation + cash",    "Parallel track → faster path to SOCOM direct contracts",           "📅 Submit Aug"],
    ["Seed",        "$8M–$10M",    "TBD",         "May 2027",     "B2B2G scale + ops cell",     "300 units B2B2G · $6.5M–$20M rev · Both SBIR Ph.I awards",        "📅 May 2027"],
    ["SBIR Ph.II",  "$750K–$2M×2","Non-dilutive","2027–2028",    "Unlocks DoD + SOCOM direct", "DoD & SOCOM Phase II → direct contracts both tracks",             "📅 2027–28"],
    ["Series A",    "$50M",        "TBD",         "May 2028",     "Direct DoD + utility scale", "SOCOM contracts · Utility pilots · $30M+ rev",                    "📅 May 2028"],
    ["Series B+",   "TBD",         "TBD",         "2029–2030",    "Market leadership + Moon",   "$200M+ rev · International · Moon roadmap",                       "📅 2029+"],
]

row_colors = ["EBF5FB","D5F5E3","FEF9E7","FDEBD0","E8DAEF",LIGHT_PURPLE,"FDEDEC","D5F5E3"]
for r, (row, bg) in enumerate(zip(rounds, row_colors), 8):
    for c_, val in enumerate(row, 1):
        cell = ws0.cell(row=r, column=c_, value=val)
        cel(cell, bold=(c_==1), bg=bg, align="center")
        bdr(cell)
    rh(ws0, r, 26)

# Revenue path
merge_hdr(ws0, 15, 1, 8, "REVENUE PATH TO $1B — BY PHASE", LIGHT_BLUE, WHITE, 12)
for i, h_ in enumerate(["Year","Phase","Units Sold","Rev: Hardware","Rev: Subscription","Rev: Operations","TOTAL Revenue","Primary Market"],1):
    c = ws0.cell(row=16, column=i, value=h_)
    hdr(c, BLUE, WHITE)
rh(ws0, 16, 28)

rev_path = [
    ["2026","Phase 1\nB2B2G only",        "12",       "$650K\n(ADAM: $80K pilot + $570K order)", "$0",       "$0",        "$650K",              "B2B2G only (ADAM, Chariot)\nSBIR submitted — no money yet\nNo direct DoD yet"],
    ["2027","Phase 1→2\nB2B2G + SBIR",   "100–300",  "$6.5M–$19.5M",  "$60K",     "$0",        "$6.5M–$20M",         "300 drones via B2B2G\nSBIR Phase I award ~March 2027\nSeed raise May 2027\nSBIR Phase II application"],
    ["2028","Phase 3\nDirect DoD + Utils","300–600",  "$19.5M–$39M",   "$1.5M",    "$5M–$10M",  "$26M–$51M",          "SOCOM + direct DoD contracts\nSBIR Phase II award\nUtility pilots begin (18mo post-Pre-Seed)\nSeries A $50M"],
    ["2029","Phase 4\nScale",             "600–1500", "$39M–$97.5M",   "$5M–$10M", "$20M–$50M", "$100M–$200M",        "DoD direct at scale\nUtility commercial rollout\nInternational expansion"],
    ["2030","Phase 4\n$1B Target",        "2000+",    "$130M+",        "$25M+",    "$200M+",    "$1,000,000,000",     "Full scale · Moon roadmap\nIPO / late stage"],
]

rev_colors = ["EBF5FB","D5F5E3","FDEBD0",LIGHT_PURPLE,"FDEDEC"]
for r, (row, bg) in enumerate(zip(rev_path, rev_colors), 17):
    bold = r == 21
    color = RED if r == 21 else "000000"
    for c_, val in enumerate(row, 1):
        cell = ws0.cell(row=r, column=c_, value=val)
        cel(cell, bold=bold, color=color, bg=bg, align="center")
        bdr(cell)
    rh(ws0, r, 24)

for col, width in [(1,8),(2,22),(3,14),(4,18),(5,16),(6,16),(7,18),(8,28)]:
    cw(ws0, col, width)


# ══════════════════════════════════════════════════════════
# SHEET 2 — BUSINESS MODEL
# ══════════════════════════════════════════════════════════
ws_bm = wb.create_sheet("Business Model")

ws_bm.merge_cells("A1:F1")
ws_bm["A1"] = "WIREBIRD — BUSINESS MODEL & PRICING"
hdr(ws_bm["A1"], NAVY, WHITE, True, 15)
rh(ws_bm, 1, 45)

# Pricing
merge_hdr(ws_bm, 3, 1, 6, "PRICING", LIGHT_BLUE, WHITE, 12)
for i, h_ in enumerate(["Revenue Stream","Price","Frequency","Market","Notes",""],1):
    c = ws_bm.cell(row=4, column=i, value=h_)
    hdr(c, BLUE, WHITE)

pricing = [
    ["Hardware — Drone Unit",    "$65,000",    "One-time per unit",  "Defense + Commercial", "Core product",                      ""],
    ["Subscription / Maintenance","$5,000",    "Per unit / per year","All clients",          "Recurring revenue — grows with fleet",""],
    ["Operations Services",      "TBD",        "Per flight / SLA",   "Utilities (Phase 2)",  "Wirebird operates drones for client",""],
    ["Pilot Programs",           "TBD",        "Project-based",      "Agriculture / Special","Energy delivery near power lines",   ""],
]

for r, row in enumerate(pricing, 5):
    bg = LIGHT_GRAY if r % 2 == 0 else WHITE
    for c_, val in enumerate(row, 1):
        cell = ws_bm.cell(row=r, column=c_, value=val)
        cel(cell, bold=(c_==1), bg=bg, align="center" if c_>1 else "left")
        bdr(cell)
    rh(ws_bm, r, 24)

# Unit economics
merge_hdr(ws_bm, 10, 1, 6, "UNIT ECONOMICS (per drone)", LIGHT_BLUE, WHITE, 12)
for i, h_ in enumerate(["Item","Amount","Notes","","",""],1):
    c = ws_bm.cell(row=11, column=i, value=h_)
    hdr(c, BLUE, WHITE)

economics = [
    ["Selling price",               "$65,000",  "One-time hardware sale"],
    ["COGS (est. components)",      "~$3,000",  "7 prototypes @ $3K — scales with volume"],
    ["Gross margin (hardware)",     "~95%",     "High margin — typical for deeptech hardware"],
    ["Annual subscription",         "$5,000",   "Per unit per year"],
    ["LTV per unit (5 years)",      "$90,000",  "$65K hardware + 5 × $5K subscription"],
]

for r, row in enumerate(economics, 12):
    bold = r in [14, 17]
    bg = LIGHT_GREEN if r == 17 else (LIGHT_GRAY if r % 2 == 0 else WHITE)
    for c_, val in enumerate(row, 1):
        cell = ws_bm.cell(row=r, column=c_, value=val if c_ <= 3 else "")
        cel(cell, bold=bold, bg=bg, align="center" if c_>1 else "left")
        bdr(cell)
    ws_bm.merge_cells(f"C{r}:F{r}")
    rh(ws_bm, r, 22)

# Go-to-market phases
merge_hdr(ws_bm, 18, 1, 6, "GO-TO-MARKET — 3 PHASES", LIGHT_BLUE, WHITE, 12)
for i, h_ in enumerate(["Phase","Timeline","Market","Business Model","Key Actions","Certification Needed"],1):
    c = ws_bm.cell(row=19, column=i, value=h_)
    hdr(c, BLUE, WHITE)

gtm = [
    ["Phase 1\nB2B2G\n(NOW)",
     "2026\nPre-Seed close",
     "ADAM · Chariot · Prime contractors\n(NOT direct DoD yet)",
     "B2B2G Hardware sales\n$65K/unit + $5K/yr\nPilot → volume deals",
     "Close ADAM deal ($80K pilot → $650K) · Chariot contract · JIFX & T-REX demos · Submit SBIR Phase I (Aug 2026)",
     "None yet — selling through primes\n⚠️ SBIR submitted but money arrives in ~9 months (March 2027)"],
    ["Phase 2\nDual SBIR + B2B2G Scale\n(2027)",
     "Early 2027\n(~9 months post-submit)",
     "DoD SBIR Phase I award\nSOCOM SBIR Phase I award\nB2B2G volume (Chariot + others)",
     "B2B2G volume + dual SBIR funding\n~$300K–$600K non-dilutive\nSBIR Phase II applications",
     "Both SBIR Phase I awards (~March 2027) · Apply both Phase II · Seed raise May 2027 · 300 units B2B2G\nTimeline: 3 months start + 2 months qualify + review = ~9 months",
     "DoD SBIR → broad DoD contracts\nSOCOM SBIR → direct SOCOM path\nRun both tracks in parallel from August 2026"],
    ["Phase 3\nDirect DoD / SOCOM\n(2028)",
     "2028\nPost-SBIR Phase II",
     "SOCOM · DoD direct contracts\nUtility pilots begin (~18 months post-Pre-Seed)",
     "Direct DoD contracts + B2B2G\nOps cell for utilities (pilot)",
     "SBIR Phase II award → direct DoD · SOCOM contract · Series A $50M · Utility pilot programs launch",
     "FAA/DoD certs accelerated via gov relationships\nUtility: 1.5–3 yrs certs but gov contacts help"],
    ["Phase 4\nScale + Space\n(2029–2030)",
     "2029–2030",
     "DoD at scale · International\nUtility commercial rollout\nMoon roadmap",
     "Full platform:\nHW + Subscription + Ops + Licensing",
     "International expansion · DoD international · Utility full rollout · Moon energy distribution R&D",
     "International airspace · ESA / NASA · Space certs"],
]

phase_colors = ["EBF5FB","FDEBD0",LIGHT_PURPLE]
for r, (row, bg) in enumerate(zip(gtm, phase_colors), 20):
    for c_, val in enumerate(row, 1):
        cell = ws_bm.cell(row=r, column=c_, value=val)
        cel(cell, bold=(c_==1), bg=bg, align="center" if c_ in [1,2] else "left")
        bdr(cell)
    rh(ws_bm, r, 60)

# Client profile
merge_hdr(ws_bm, 24, 1, 6, "CLIENT PROFILE", LIGHT_BLUE, WHITE, 12)
for i, h_ in enumerate(["Segment","Client Size","Deal Size","Timeline","# Clients Needed (2027)","Notes"],1):
    c = ws_bm.cell(row=25, column=i, value=h_)
    hdr(c, BLUE, WHITE)

clients = [
    ["🔴 ADAM (active)",  "First client",           "$650K total", "NOW → post-JIFX","1 client / 12 units","$80K pilot (2 units) → 10 units @ $65K post-JIFX. $80K deducted from order."],
    ["DoD Direct",        "US Gov. agencies",       "$500K–$5M+",  "2026–2027","3–5 contracts",     "JIFX + T-REX + SBIR path"],
    ["B2B2G Partners",    "Prime contractors",      "$200K–$2M",   "2027",     "5–10 partners",     "Anduril-type relationships"],
    ["Utilities (Phase 2)","Top 10 US utilities",   "$1M–$10M/yr", "2028",     "2–3 utility pilots","Operations cell model"],
    ["Ag / Special Pilots","SMBs near power lines", "$50K–$200K",  "2028+",    "TBD",               "Energy delivery near lines"],
]

for r, row in enumerate(clients, 26):
    bg = LIGHT_GRAY if r % 2 == 0 else WHITE
    for c_, val in enumerate(row, 1):
        cell = ws_bm.cell(row=r, column=c_, value=val)
        cel(cell, bold=(c_==1), bg=bg, align="center" if c_>1 else "left")
        bdr(cell)
    rh(ws_bm, r, 24)

# Competitors / benchmark
merge_hdr(ws_bm, 31, 1, 6, "BENCHMARK", LIGHT_BLUE, WHITE, 12)
for i, h_ in enumerate(["Company","Raised","Model","Relevance",""],1):
    c = ws_bm.cell(row=32, column=i, value=h_)
    hdr(c, BLUE, WHITE)
ws_bm.merge_cells("E32:F32")

benchmarks = [
    ["WindLift",  "$20M",  "Airborne wind energy",          "Similar deeptech raise trajectory — target comparable raise"],
    ["Anduril",   "$4.6B", "Defense autonomy platform",     "HQ neighbor in LA — ecosystem + potential partner"],
    ["SpaceX",    "—",     "Space + energy infrastructure", "Long-term vision alignment — LA ecosystem"],
    ["Wirebird",  "—",     "Power line energy distribution","White space: no direct competitor confirmed"],
]

for r, row in enumerate(benchmarks, 33):
    bg = LIGHT_GREEN if r == 36 else (LIGHT_GRAY if r % 2 == 0 else WHITE)
    for c_, val in enumerate(row, 1):
        cell = ws_bm.cell(row=r, column=c_, value=val if c_ <= 4 else "")
        cel(cell, bold=(c_==1 or r==36), bg=bg, align="center" if c_ in [1,2] else "left")
        bdr(cell)
    ws_bm.merge_cells(f"D{r}:F{r}")
    rh(ws_bm, r, 22)

for col, width in [(1,18),(2,18),(3,28),(4,38),(5,28),(6,14)]:
    cw(ws_bm, col, width)


# ══════════════════════════════════════════════════════════
# SHEET 3 — 6-MONTH BUDGET
# ══════════════════════════════════════════════════════════
ws1 = wb.create_sheet("6-Month Budget")

months = ["June 2026","July 2026","Aug 2026","Sep 2026","Oct 2026","Nov 2026"]

ws1.merge_cells("A1:I1")
ws1["A1"] = "WIREBIRD — 6-MONTH EXPENSE PLAN (Jun–Nov 2026)"
hdr(ws1["A1"], NAVY, WHITE, True, 15)
rh(ws1, 1, 45)

ws1.merge_cells("A2:I2")
ws1["A2"] = "All amounts in USD · Bridge target: $200K · Pre-Seed close: October 2026"
hdr(ws1["A2"], DARK_GRAY, WHITE, False, 9)
rh(ws1, 2, 18)

headers = ["Category","Item"] + months + ["6-Month Total"]
for c, val in enumerate(headers, 1):
    cell = ws1.cell(row=3, column=c, value=val)
    hdr(cell, BLUE, WHITE)
rh(ws1, 3, 28)

expense_rows = [
    ("TEAM",          "Anis Cheriet — CEO (49%)",              7000, 7000, 7000, 7000, 7000, 7000),
    ("TEAM",          "Bo — CTO (51%)",                        7000, 7000, 7000, 7000, 7000, 7000),
    ("TEAM",          "Armin (Part-time)",                     3500, 3500, 3500, 3500, 3500, 3500),
    ("TEAM",          "Becket (Consultant)",                   3500, 3500, 3500, 3500, 3500, 3500),
    ("HOUSING",       "Founders Inc (included in program)",       0,    0,    0,    0,    0,    0),
    ("HOUSING",       "House + Garage 4BR (workspace)",           0,    0, 7000, 7000, 7000, 7000),
    ("HARDWARE",      "Drone prototypes ×7 @ $3K",            6000, 3000, 3000, 3000, 3000, 3000),
    ("HARDWARE",      "Lab & bench equipment (one-time)",      2200,    0,    0,    0,    0,    0),
    ("HARDWARE",      "Carbon fiber & raw materials",           500,  500,  500,  500,  500,  500),
    ("EVENTS",        "JIFX — travel + shipping + hotel",         0, 2000, 8000,    0,    0,    0),
    ("EVENTS",        "T-REX 26-3 — travel + logistics",          0,    0,    0,20000,    0,    0),
    ("EVENTS",        "Event #3 (TBD)",                           0,    0,    0,    0, 5000,    0),
    ("EVENTS",        "Event #4 (TBD)",                           0,    0,    0,    0,    0, 5000),
    ("LEGAL & IP",    "Provisional patent — USPTO (~$1.5K)",      0, 2000,    0,    0,    0,    0),
    ("LEGAL & IP",    "Patent attorney fees",                     0,  500,  500,  500,    0,    0),
    ("LEGAL & IP",    "Clerky (SAFE documents)",               200,    0,    0,    0,    0,    0),
    ("SUBSCRIPTIONS", "AI tools (Claude, GPT, Cursor, etc.)",   300,  300,  300,  300,  300,  300),
    ("SUBSCRIPTIONS", "CAD / simulation software",              200,  200,  200,  200,  200,  200),
    ("SUBSCRIPTIONS", "Other SaaS",                            100,  100,  100,  100,  100,  100),
    ("MISC",          "Transportation & logistics",             300,  300,  300,  300,  300,  300),
    ("MISC",          "Buffer / unexpected",                   1000, 1000, 1000, 1000, 1000, 1000),
]

cat_colors = {
    "TEAM":"EBF5FB","HOUSING":"E9F7EF","HARDWARE":"FEF9E7",
    "EVENTS":"FDEDEC","LEGAL & IP":LIGHT_PURPLE,
    "SUBSCRIPTIONS":"EBF5FB","MISC":LIGHT_GRAY,
}

current_row = 4
current_cat = None
monthly_totals = [0]*6

for item_row in expense_rows:
    cat, item = item_row[0], item_row[1]
    values = list(item_row[2:])
    total = sum(values)
    bg = cat_colors.get(cat, WHITE)

    if cat != current_cat:
        current_cat = cat
        ws1.merge_cells(f"A{current_row}:I{current_row}")
        c = ws1.cell(row=current_row, column=1, value=cat)
        hdr(c, LIGHT_BLUE, WHITE, True, 10, "left")
        rh(ws1, current_row, 20)
        current_row += 1

    ws1.cell(row=current_row, column=1).fill = PatternFill("solid", fgColor=bg)
    cell2 = ws1.cell(row=current_row, column=2, value=item)
    cel(cell2, bg=bg)
    bdr(cell2)

    for ci, val in enumerate(values, 3):
        monthly_totals[ci-3] += val
        c = ws1.cell(row=current_row, column=ci, value=val if val else "—")
        cel(c, bg=bg if val else WHITE, align="center")
        if isinstance(val, int) and val > 0:
            c.number_format = '$#,##0'
        bdr(c)

    tc = ws1.cell(row=current_row, column=9, value=total)
    cel(tc, bold=True, bg=bg, align="center")
    tc.number_format = '$#,##0'
    bdr(tc)
    rh(ws1, current_row, 20)
    current_row += 1

grand_total = sum(monthly_totals)
ws1.cell(row=current_row, column=1, value="MONTHLY TOTAL")
ws1.merge_cells(f"A{current_row}:B{current_row}")
hdr(ws1.cell(row=current_row, column=1), GREEN, WHITE, True, 11, "center")
for ci, t in enumerate(monthly_totals, 3):
    c = ws1.cell(row=current_row, column=ci, value=t)
    hdr(c, GREEN, WHITE, True, 11)
    c.number_format = '$#,##0'
gc = ws1.cell(row=current_row, column=9, value=grand_total)
hdr(gc, GREEN, WHITE, True, 12)
gc.number_format = '$#,##0'
rh(ws1, current_row, 30)

current_row += 2
ws1.merge_cells(f"A{current_row}:I{current_row}")
note = ws1.cell(row=current_row, column=1,
    value=f"✅  6-Month Total: ${grand_total:,.0f}  |  Bridge: $200K  |  Gap: covered by client $80K revenue + SBIR Phase I ($150K–$300K non-dilutive)")
note.font = Font(bold=True, color=LIGHT_BLUE, size=10, name="Calibri")
note.fill = PatternFill("solid", fgColor="EBF5FB")
rh(ws1, current_row, 22)

for col, width in [(1,14),(2,36),(3,13),(4,13),(5,13),(6,13),(7,13),(8,13),(9,15)]:
    cw(ws1, col, width)


# ══════════════════════════════════════════════════════════
# SHEET 4 — CAP TABLE
# ══════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Cap Table")

ws2.merge_cells("A1:E1")
ws2["A1"] = "CAP TABLE SIMULATION"
hdr(ws2["A1"], NAVY, WHITE, True, 14)
rh(ws2, 1, 35)

merge_hdr(ws2, 3, 1, 5, "TODAY (Pre-funding)", DARK_GRAY, WHITE, 11)
for i, h_ in enumerate(["Shareholder","Role","% Ownership","Type","Notes"],1):
    c = ws2.cell(row=4, column=i, value=h_)
    hdr(c, BLUE, WHITE)

today = [
    ["Anis Cheriet",  "CEO & Co-founder",  "49%", "Common shares",   "Co-founder"],
    ["Bo",            "CTO & Co-founder",  "51%", "Common shares",   "Co-founder"],
    ["Armin",         "Part-time",         "TBD", "TBD / Cash",      "⚠️ Clarify before Pre-Seed"],
    ["Becket",        "Consultant",        "TBD", "TBD / Cash",      "⚠️ Clarify before Pre-Seed"],
    ["TOTAL",         "",                  "100%","",                "Pre-funding"],
]
for r, row in enumerate(today, 5):
    bold = r == 9
    bg = LIGHT_GRAY if r == 9 else ("FFF3CD" if r in [7,8] else WHITE)
    for c_, val in enumerate(row, 1):
        cell = ws2.cell(row=r, column=c_, value=val)
        cel(cell, bold=bold, bg=bg, align="center")
        bdr(cell)

ws2.merge_cells("A11:E11")
ws2["A11"] = "⚠️  Armin & Becket equity TBD — investors will ask. Clarify now: equity grant vs. cash-only."
ws2["A11"].font = Font(italic=True, color=ORANGE, size=10, name="Calibri")
ws2["A11"].fill = PatternFill("solid", fgColor=YELLOW)

merge_hdr(ws2, 13, 1, 5, "AFTER PRE-SEED ($2M @ $10–15M valuation) — SAFE converts", GREEN, WHITE, 11)
for i, h_ in enumerate(["Shareholder","Amount Invested","% (approx)","Valuation Used","Notes"],1):
    c = ws2.cell(row=14, column=i, value=h_)
    hdr(c, BLUE, WHITE)

preseed = [
    ["Anis Cheriet",         "—",          "~38%", "—",          "Founder"],
    ["Bo",                   "—",          "~39%", "—",          "Founder"],
    ["Angels (SAFE $200K)",  "$200,000",   "~3%",  "$5M cap",    "Converts at Pre-Seed"],
    ["Pre-Seed VCs ($2M)",   "$2,000,000", "~15%", "$10–15M",    "New round"],
    ["Option Pool (ESOP)",   "—",          "~10%", "—",          "Create before close"],
    ["TOTAL",                "$2,200,000", "~105%","",           "* approx"],
]
for r, row in enumerate(preseed, 15):
    bold = r == 20
    bg = LIGHT_GREEN if r in [15,16] else (LIGHT_GRAY if r==20 else WHITE)
    for c_, val in enumerate(row, 1):
        cell = ws2.cell(row=r, column=c_, value=val)
        cel(cell, bold=bold, bg=bg, align="center")
        bdr(cell)

for col, width in [(1,24),(2,18),(3,14),(4,18),(5,35)]:
    cw(ws2, col, width)


# ══════════════════════════════════════════════════════════
# SHEET 5 — MILESTONES
# ══════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Milestones")

ws3.merge_cells("A1:F1")
ws3["A1"] = "RETROGRADE PLAN — PRE-SEED CLOSE OCTOBER 2026"
hdr(ws3["A1"], NAVY, WHITE, True, 14)
rh(ws3, 1, 35)

for i, h_ in enumerate(["Phase","Deadline","Action","Priority","Impact","Status"],1):
    c = ws3.cell(row=2, column=i, value=h_)
    hdr(c, BLUE, WHITE)

phase_colors = {
    "OCTOBER":"D5F5E3","SEPTEMBER":"D6EAF8",
    "AUGUST":"FEF9E7","JULY":"FDEBD0","JUNE":"FDEDEC"
}

milestones = [
    ["OCTOBER",   "Oct 2026",  "Close Pre-Seed $2M @ $10–15M valuation",        "🎯","Production funding",    "⏳ Target"],
    ["SEPTEMBER", "Sep 2026",  "15–20 VC meetings (dual-use / defense)",         "🔴","Close the round",       "📅 Upcoming"],
    ["SEPTEMBER", "Sep 2026",  "2–3 term sheets received",                       "🔴","Negotiate terms",       "📅 Upcoming"],
    ["SEPTEMBER", "Sep 2026",  "Pre-seed deck finalized & rehearsed",            "🔴","Fundraising tool",      "📅 Upcoming"],
    ["SEPTEMBER", "Sep 2026",  "2–3 LOIs from clients (DoD / utilities)",        "🔴","Proof of demand",       "📅 Post-events"],
    ["AUGUST",    "Aug 2026",  "JIFX demo — TRL 4 → TRL 5",                     "🔴","Field validation",      "📅 In 2 months"],
    ["AUGUST",    "Aug 2026",  "SBIR Phase I submitted",                         "🔴","$150K–300K non-dilut.", "📅 To do"],
    ["AUGUST",    "Aug 2026",  "Bridge $200K closed",                            "🔴","Runway to October",     "📅 Launch now"],
    ["AUGUST",    "Aug 2026",  "T-REX 26-3 — participate if slot obtained",     "🟡","DoD credibility",       "📅 Apply now"],
    ["JULY",      "Jul 15",    "Provisional patent filed (USPTO) — ~$1,500",    "🔴","IP before VCs",         "📅 To do"],
    ["JULY",      "Jul 1",     "Client $40K signed (Chariot deal)",             "🔴","Commercial proof",      "🟡 In progress"],
    ["JULY",      "Jul 2026",  "Bridge SAFE rolled to angels",                  "🔴","Fund operations",       "📅 To do"],
    ["JULY",      "Jul 2026",  "Confirm Armin & Becket equity or cash",         "🟡","Clean cap table",       "📅 To do"],
    ["JUNE",      "This week", "Apply T-REX 26-3 → trex.events",               "🔴 URGENT","DoD access",     "🔴 Today"],
    ["JUNE",      "This week", "Create SAFE $20K on Clerky ($5M cap)",         "🔴 URGENT","Bridge now",     "🔴 Today"],
    ["JUNE",      "This week", "Contact patent attorney (provisional)",         "🔴 URGENT","Protect IP",     "🔴 Today"],
]

for r, row in enumerate(milestones, 3):
    bg = phase_colors.get(row[0], WHITE)
    for c_, val in enumerate(row, 1):
        cell = ws3.cell(row=r, column=c_, value=val)
        cel(cell, bg=bg, align="center")
        bdr(cell)
    rh(ws3, r, 22)

for col, width in [(1,13),(2,13),(3,44),(4,14),(5,22),(6,16)]:
    cw(ws3, col, width)


# ══════════════════════════════════════════════════════════
# SHEET 6 — BRIDGE TRACKER
# ══════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Bridge Tracker")

ws4.merge_cells("A1:E1")
ws4["A1"] = "BRIDGE ROUND TRACKER — SAFE @ $5M CAP"
hdr(ws4["A1"], NAVY, WHITE, True, 14)
rh(ws4, 1, 35)

for r, (label, val, bg) in enumerate([
    ("TARGET","$200,000",LIGHT_GRAY),
    ("CLOSED TO DATE","$0",PINK),
    ("REMAINING","$200,000",YELLOW),
], 3):
    ws4.cell(row=r,column=1,value=label).font = Font(bold=True,name="Calibri",size=10)
    ws4.cell(row=r,column=1).fill = PatternFill("solid",fgColor=bg)
    ws4.merge_cells(f"A{r}:D{r}")
    v = ws4.cell(row=r,column=5,value=val)
    v.font = Font(bold=True,name="Calibri",size=11)
    v.alignment = Alignment(horizontal="center",vertical="center")
    v.fill = PatternFill("solid",fgColor=bg)

for i, h_ in enumerate(["Investor","Amount","Status","SAFE Signed Date","Notes"],1):
    c = ws4.cell(row=7,column=i,value=h_)
    hdr(c, BLUE, WHITE)

angels = [
    ["Angel 1","$40,000","🔴 In progress","","First check — Clerky SAFE this week"],
    ["Angel 2","$50,000","📅 To approach","",""],
    ["Angel 3","$50,000","📅 To approach","",""],
    ["Angel 4","$60,000","📅 To approach","",""],
    ["TOTAL","$200,000","","",""],
]
for r, row in enumerate(angels, 8):
    bold = r == 12
    bg = LIGHT_GRAY if r == 12 else WHITE
    for c_, val in enumerate(row, 1):
        cell = ws4.cell(row=r,column=c_,value=val)
        cel(cell,bold=bold,bg=bg,align="center")
        bdr(cell)

ws4.merge_cells("A14:E14")
n = ws4.cell(row=14,column=1,
    value="💡  If bridge doesn't close fully → SBIR Phase I ($150K–$300K non-dilutive). Submit by August.")
n.font = Font(italic=True,color=LIGHT_BLUE,size=10,name="Calibri")
n.fill = PatternFill("solid",fgColor="EBF5FB")

for col, width in [(1,20),(2,14),(3,18),(4,18),(5,38)]:
    cw(ws4, col, width)


# ══════════════════════════════════════════════════════════
# SAVE
# ══════════════════════════════════════════════════════════
import os
path = os.path.expanduser("~/Desktop/Wirebird_Fundraising_Strategy.xlsx")
wb.save(path)
print(f"✅  Saved: {path}")
print(f"    6-month total: ${grand_total:,.0f}")
print(f"    Monthly: {[f'${t:,}' for t in monthly_totals]}")
